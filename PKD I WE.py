#LIBRERÍAS
import pandas as pd
import numpy as np
import random
import re
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import datetime
import time
from datetime import timedelta
import calendar
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
import urllib3
import shutil
import locale


def codigo_completo ():
    intento = 0
    referencia = 3  

    while intento < 3:
        try:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')
            carpeta = os.path.join(os.environ['USERPROFILE'], 'OneDrive - Truper, S.A. de C.V', 'Escritorio', 'Análisis MeLi')
            os.makedirs(carpeta, exist_ok = True)
            screenshot_folder = os.path.join(carpeta, '3. Screenshots MN')
            try:
                shutil.rmtree(screenshot_folder)
            except:
                pass
            os.makedirs(screenshot_folder, exist_ok=True)

            ##Buscar productos 
            def buscar_productos(driver, url, pattern):
                
                
                driver.get(url)
                time.sleep(random.uniform(3.5, 6.5))
                
            ##Busca las cards versión andes y poly
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.andes-card'))        )
                    encontrados = 'div.andes-card'
                except:
                    print('No es una página andes')
                    try:
                        WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.poly-card'))        )
                        encontrados = 'div.poly-card'
                    except:
                        try:
                            WebDriverWait(driver, 10).until(
                                EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.andes-card'))        )
                            encontrados = 'div.andes-card'
                        except:
                            print('No es una página andes')
                            try:
                                WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.poly-card'))        )
                                encontrados = 'div.poly-card'
                                print('No hay productos')
                            except:
                                return [],[],[],[],[]

            ##Asigna las cards a productos
                productos = driver.find_elements(By.CSS_SELECTOR, encontrados)
                links, mlms, titulos, prices, tiendas = [], [], [], [], []
                
                for producto in productos:
                    #LIKS Y MLMs
                    try:
                        link = producto.find_element(By.CSS_SELECTOR, 'a').get_attribute('href')
                        links.append(link)
                        
                        match = re.search(pattern, link)
                        mlms.append(match.group().replace('-', '') if match else 'MLM')
                    except:
                        links.append(np.nan)
                        mlms.append(np.nan)
                    
                    #DESCRIPCIONES
                    #time.sleep(random.uniform(0.1, 0.5))
                    try:
                        title = producto.find_element(By.CSS_SELECTOR, 'a').text
                        titulos.append(title)
                    except:
                        titulos.append(np.nan)
                   
                    #PRECIOS
                    try:
                        price = producto.find_element(By.CSS_SELECTOR, 'div.poly-price__current span.andes-money-amount')
                        price = price.get_attribute('aria-label').strip()
                        partes = price.split()
                        numeros = [p for p in partes if p.isdigit()]
                        if len(numeros) >= 2:
                            entero = int(numeros[0])
                            centavos = int(numeros[1])
                            price = float(f'{entero}.{centavos:02d}')
                        else:
                            price = float(f'{numeros[0]}.00')
                        prices.append(price)
                    except:
                        prices.append(np.nan)
                    
                        #TIENDA OFICIAL
                    try:
                        tienda = producto.find_element(By.CSS_SELECTOR, 'span.poly-component__seller')
                        tienda = tienda.text.replace('Por ','').strip()
                        tiendas.append(tienda)
                    except:
                        tiendas.append(np.nan)
                
                return links, mlms, titulos, prices, tiendas

            ## Configuración de Driver
            def configurar_driver():
                chrome_options = Options()
                chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36')
                chrome_options.add_argument('--window-size=1920x1080')
                path_driver = os.path.join(carpeta, 'ChromeDriver1', 'chromedriver.exe')
                service = Service(path_driver)
                driver = webdriver.Chrome(service=service, options=chrome_options)
                driver.set_page_load_timeout(60)
                return driver


            def login_meli(driver):
               
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                from selenium.webdriver.common.by import By
                import time
           
                wait = WebDriverWait(driver, 30)
                url_login = "https://www.mercadolibre.com/jms/mlm/lgz/login?platform_id=ml&go=https://listado.mercadolibre.com.mx/ROTO-1%2F2A8-Truper?sb=all_mercadolibre&loginType=negative_traffic"
                driver.get(url_login)
           
                for intento in range(2):
                    try:
                        time.sleep(6)
                        email_input = wait.until(EC.presence_of_element_located((By.ID, "user_id")))
                        email_input.clear()
                        email_input.send_keys("dreynosoh@truper.com")
           
                        time.sleep(2)
                        continuar_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[span[text()="Continuar"]]')))
                        continuar_btn.click()
                        print(f"✅ Intento {intento+1}: correo ingresado y 'Continuar' presionado")
                        time.sleep(10)
                        break
                    except Exception as e:
                        print(f"⚠️ Falló intento {intento+1}: {e}")
                        time.sleep(8)
           
                try:
                    captcha_iframe = wait.until(EC.presence_of_element_located(
                        (By.XPATH, '//iframe[contains(@src, "recaptcha")]')
                    ))
                    driver.switch_to.frame(captcha_iframe)
           
                    checkbox = wait.until(EC.element_to_be_clickable((By.ID, "recaptcha-anchor")))
                    checkbox.click()
                    print("🧠 reCAPTCHA clickeado")
                    time.sleep(8)
           
                    driver.switch_to.default_content()
                except Exception:
                    print("🔕 No apareció reCAPTCHA o no fue necesario")
           
                time.sleep(8)
           
                try:
                    continuar_btn2 = wait.until(EC.element_to_be_clickable((By.XPATH, '//button[span[text()="Continuar"]]')))
                    continuar_btn2.click()
                    print("🔄 Segundo clic en 'Continuar'")
                    time.sleep(10)
                except Exception as e:
                    print(f"⚠️ No se pudo dar el segundo clic en 'Continuar': {e}")
           
                try:
                    email_option_btn = wait.until(EC.element_to_be_clickable((
                        By.XPATH, '//button[.//span[text()="E-mail"]]'
                    )))
                    email_option_btn.click()
                    print("📩 Opción 'E-mail' seleccionada")
                    time.sleep(50)
                except Exception as e:
                    print(f"❌ No se pudo hacer clic en la opción 'E-mail': {e}")


            #Inicio de ejecución del código
            start_time = time.time()

            # BASE DE DATOS
            file = r'C:\Users\dreynosoh\Downloads\Base de Datos 3.xlsx'
            df = pd.read_excel(file)
            dfs = []
            pattern = r'MLM[-]?\d+' #'item_id=\d+'
            pattern2 = r'(seller_id|official_store_id)=\d+'
            horas, vendedores, ids = [], [], []
            driver = configurar_driver()
            login_meli(driver)

            try:
                for i, rowOG in df.iloc[referencia:4].iterrows():
                    print(f'🔍 Búsqueda {i+1} de {len(df)} - Código: {rowOG["Código"]}, Clave: {rowOG["Clave"]}')
                    
                    links_cod, mlm_cod, descs_cod, prec_cod, tie_cod = buscar_productos(driver, rowOG['Link Código'], pattern)
                    links_clav, mlm_clav, descs_clav, prec_clav, tie_clav = buscar_productos(driver, rowOG['Link Clave'], pattern)
                    
                    
                    total_links = links_cod + links_clav

                    size = len(total_links)

                    aux_df = pd.DataFrame({
                        'Código Truper': [rowOG['Código']] * size,
                        'Clave Truper': [rowOG['Clave']] * size,
                        'Descripción Truper': [rowOG['Título del Producto']] * size,
                        'Descripción de la Publicación': descs_cod + descs_clav,
                        'Id': mlm_cod + mlm_clav,
                        'Precio Mayoreo': [rowOG['Precio FINAL']] * size,
                        'Precio MeLi': prec_cod + prec_clav,
                        'T. Oficial': tie_cod + tie_clav,
                        'Link': total_links
                    })
                    aux_df = aux_df[~aux_df['Link'].str.startswith('https://click1.mercadolibre.com.mx/brand_ads', na = False)]
                    aux_df = aux_df[~aux_df['Link'].str.startswith('https://www.mercadolibre.com.mx/blog', na = False)]
                    
                    referencia = i
                    for idx, row in aux_df[aux_df['Id'] == 'MLM'].iterrows():
                        if 'truper' in row['Link']:
                            driver.get(row['Link'])
                      
                        time.sleep(random.uniform(3.5, 6.5))
                        redirected_url = driver.current_url
                
                        match = re.search(pattern, redirected_url)
                        if match:
                            new_id = match.group().replace('-', '')
                            aux_df.at[idx, 'Link'] = redirected_url
                            aux_df.at[idx, 'Id'] = new_id
                            
                    for idx, row in aux_df[aux_df['Descripción de la Publicación'].isna()].iterrows():
                        driver.get(row['Link'])
                        time.sleep(random.uniform(3.5, 6.5))
                        newtitulo = driver.find_elements(By.CLASS_NAME, "ui-pdp-title")[0].text
                        aux_df.at[idx, 'Descripción de la Publicación'] = newtitulo
                    
                    #AQUÍ ESTÁ UN CHANCE DE MEJORAR PARA ASEGURAR QUE TODAS LAS DESCRIPCIONES SE EXTRAIGAN
                    #SI LO LOGRAS PUEDES ELIMINAR EL dropna()
                    
                    #Inicio del análisis
                    aux_df = aux_df.dropna()
                    aux_df = aux_df.drop_duplicates(subset = ['Id'], keep = 'first')
                    
                    #Primer filtro: Diferencia de precios (Para quitar refacciones)
                    p = rowOG['Precio FINAL']
                    db = aux_df.copy()
                    db['Diferencia'] = (db['Precio MeLi'] - p) / p
                    db_filtro1 = db[db['Diferencia'] > -0.4]
                    db_filtro1 = db_filtro1.copy()
                    db_filtro1 = db_filtro1[~db_filtro1['Descripción de la Publicación'].str.contains(r'\bpara\b', case = False, na = False, regex = True)]
                    
                    #Segundo filtro: Análisis de similitudes entre descripciones (Procesamiento de Lenguaje Natural)
                    keys = rowOG['Key Words']
                    keys = '|'.join(keys.split('; '))
                    COD = rowOG['Código']
                    CLAV = rowOG['Clave']
                    db_filtro1 = db_filtro1.copy()
                    db_filtro2 = db_filtro1[db_filtro1['Descripción de la Publicación'].str.contains(keys, case = False, na = False)]
                    db_filtro2 = db_filtro2.copy()
                    db_filtro2 = db_filtro2[db_filtro2['Descripción de la Publicación'].str.contains(fr'(?<!\S){COD}(?![\w/-])|(?<!\S){CLAV}(?![\w/-])', case=False, na=False, regex=True)]
                    
                    db_filtro2 = db_filtro2.copy()
                    db_filtro2 = db_filtro2.drop(['Diferencia'], axis = 1)
                    
                    def jaccard_similarity(text1, text2):
                        set1, set2 = set(text1.lower().split()), set(text2.lower().split())
                        return len(set1 & set2) / len(set1 | set2)
                    
                    def count_matches(text, words):
                        return sum(1 for word in words if word.lower() in text.lower())
                    
                    db_filtro2 = db_filtro2.copy()
                    JACCARD = []
                    SCIKITLEARN = []
                    matches = []
                    for trup, meli in zip(db_filtro2['Descripción Truper'], db_filtro2['Descripción de la Publicación']):
                        similarity = jaccard_similarity(trup, meli)
                        JACCARD.append(round(similarity,3))
                        
                        texts = [trup, meli]
                        vectorizer = TfidfVectorizer().fit_transform(texts)
                        similarity = cosine_similarity(vectorizer[0], vectorizer[1])[0][0]
                        SCIKITLEARN.append(round(similarity,3)) 
                        
                        KEYS = rowOG['Key Words']
                        lista = KEYS.split("; ") + [CLAV]
                        matches.append(count_matches(meli, lista))
                        
                    db_filtro2['JACCARD'] = JACCARD
                    db_filtro2['SCIKIT-LEARN'] = SCIKITLEARN
                    db_filtro2['MATCH_COUNT'] = matches
                    
                    db_filtro2 = db_filtro2.copy()
                    db_filtro2 = db_filtro2[db_filtro2['SCIKIT-LEARN'] >=  0.05]
                    db_filtro2 = db_filtro2.copy()
                    db_filtro2 = db_filtro2.loc[(db_filtro2['MATCH_COUNT'] >= 2) | ((db_filtro2['MATCH_COUNT'] <= 3) & (db_filtro2['JACCARD'] > 0.07))]
                    db_filtro2 = db_filtro2.copy()
                    db_filtro2 = db_filtro2[db_filtro2['JACCARD'] >=  0.07]
                    
                    #Tercer filtro: Menor precio (Solo nos interesan los negativos)
                    db_filtro2 = db_filtro2.copy()
                    db_filtro2['Factor'] = db_filtro2['Precio MeLi'] - p
                    db_filtro3 = db_filtro2[db_filtro2['Factor'] < 0]
                    
                    db_filtro1 = db_filtro1.copy()
                    db_filtro1 = db_filtro1.drop(['Diferencia'], axis = 1)   
                    db_filtro2 = db_filtro2.copy()
                    db_filtro2 = db_filtro2.drop(['JACCARD', 'SCIKIT-LEARN', 'MATCH_COUNT', 'Factor'], axis = 1)
                    db_warning = db_filtro3.copy()
                    db_warning = db_warning.drop(['JACCARD', 'SCIKIT-LEARN', 'MATCH_COUNT', 'Factor'], axis = 1)
                    
                    if len(db_warning) > 0:
                        links = db_warning['Link']
                        mlms = db_warning['Id']
                        prcs = db_warning['Precio MeLi']
                        
                        for mlm, link in zip(mlms, links):
                            driver.get(link)
                            time.sleep(random.uniform(3.5, 6.5))
                            
                            #Vendedores

                            # Lista donde guardaremos los vendedores
                            vendedores = []
                            
                            try:
                                # Buscar elementos de vendedores
                                sellers = driver.find_elements(By.CLASS_NAME, "ui-seller-data-header__title")
                                
                                if sellers:  # ✅ Si se encontraron vendedores, los procesamos
                                    vendedores.append(", ".join([seller.text.replace('Vendido por ', '').strip() for seller in sellers]))
                                else:
                                    vendedores.append(pd.NA)  # Si no hay vendedores, agregamos NaN
                            except:  
                                vendedores.append(pd.NA)
                            
                            # 📊 Verificar tamaños antes de la inserción
                            #print(f"🔍 Tamaño de db_warning: {len(db_warning)}")
                            #print(f"🔍 Tamaño de vendedores: {len(vendedores)}")
                            
                            # 🏗️ Ajuste del tamaño de vendedores antes de insertar la columna
                            while len(vendedores) < len(db_warning):
                                vendedores.append(pd.NA)  # Rellenar con NaN si hay menos datos
                            
                            if len(vendedores) > len(db_warning):
                                vendedores = vendedores[:len(db_warning)]  # Truncar si hay más vendedores
                            
                            # ✍ **Eliminar la columna si ya existe antes de insertar**
                            if "Nombre del Seller" in db_warning.columns:
                                db_warning.drop(columns=["Nombre del Seller"], inplace=True)
                            
                            # 🎯 Insertar la columna corregida en el DataFrame
                            db_warning["Nombre del Seller"] = pd.Series(vendedores).reindex(db_warning.index)
                            
                                            
                            #IdSellers
                            try:
                                seller_link = driver.find_element(By.CSS_SELECTOR, "div.ui-seller-data-footer__container a")
                                seller_link = seller_link.get_attribute("href")
                                match2 = re.search(pattern2, seller_link)
                                ids.append(match2.group().split('=')[1] if match2 else np.nan)
                            except:
                                ids.append(np.nan)
                            
                            screenshot_path = os.path.join(screenshot_folder, f"{mlm}.png")
                            while not os.path.exists(screenshot_path):
                                driver.execute_script("document.body.style.zoom='60%'")
                                time.sleep(random.uniform(3.5, 6.5))
                                driver.save_screenshot(screenshot_path)
                                print('Se guardó una SS 📷')
                                driver.get(link)
                            horas.append(datetime.datetime.now().strftime("%d/%m/%Y %H:%M"))
                        
                        #db_warning.insert(8, 'Nombre del Seller', vendedores)
                        dfs.append(db_warning)
                        
            except KeyboardInterrupt:
               print("🚨 Interrupción detectada! Finalizando...")     
               
            finally:
                hoy = datetime.date.today()
                mes = hoy.month
                año = hoy.year
                nombre_mes = calendar.month_name[mes].capitalize()
                mesFolder = f'{mes}. {nombre_mes} {año}'
                destino = os.path.join(carpeta, '1. Warnings', mesFolder)
                os.makedirs(destino, exist_ok = True)
                fecha = datetime.datetime.now().strftime('%Y%m%d')
                file_warnings = os.path.join(destino, f'Warnings_{fecha}.xlsx')
                print(f"Total de DataFrames en dfs: {len(dfs)}")
                print(f"DataFrames vacíos en dfs: {sum(df.empty for df in dfs)}")
                df_final = pd.concat([df for df in dfs if not df.empty], ignore_index = True)
                df_final = df_final.copy()
                
                df_final.insert(7, 'Fecha Scraping', horas)
                df_final.insert(8, 'IdSeller', ids)
                df_final.insert(0, '#', np.arange(1, len(df_final)+1))
                df_final.to_excel(file_warnings, index = False)
                links = df_final['Link']
                df_final = df_final.drop(['Link'], axis = 1)
                
                with pd.ExcelWriter(file_warnings, engine='openpyxl') as writer:
                    df_final.to_excel(writer, index=False, sheet_name='Warnings')
                
                workbook = load_workbook(file_warnings)
                workbook.create_sheet('Evidencias')
                worksheet = workbook['Warnings']
                
                column_to_exclude = worksheet['D']
                for column in worksheet.columns:
                    column_letter = column[0].column_letter
                    max_length = 0
                    for cell in column:  
                        if cell not in column_to_exclude:
                            cell.alignment = Alignment(horizontal = 'center')
                            
                        if column_letter in ['D']:
                            cell.alignment = Alignment(horizontal = 'left')
                            
                        if column_letter in ['E']:
                            cell.alignment = Alignment(horizontal = 'left')
                            
                        if column_letter in ['G']:
                            cell.number_format = '"$"#,##0'
                            
                        if column_letter in ['H']:
                            cell.number_format = '"$"#,##0.00'
                
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                for cell in worksheet[1]:
                    cell.font = Font(color = 'FFFFFF', bold = False)
                    cell.fill = PatternFill(start_color = '000000', end_color = '000000', fill_type = 'solid')
                    cell.alignment = Alignment(horizontal = 'center')
                
                for index, (mlm, link) in enumerate(zip(df_final['Id'], links), start=2):
                    mlm_cell = worksheet[f'F{index}']
                    mlm_cell.value = mlm
                    mlm_cell.hyperlink = link
                
                worksheet = workbook['Evidencias']
                worksheet['A1'] = '#'
                worksheet['B1'] = 'MLM'
                worksheet['D1'] = 'Descripción Truper'
                worksheet['F1'] = 'Precio Mayoreo'
                worksheet['H1'] = 'Precio MeLi'
                worksheet['J1'] = 'Fecha de Scraping'    
                for index, (mlm, link, hora, precML, precT, desc) in enumerate(zip(df_final['Id'], links, df_final['Fecha Scraping'], df_final['Precio MeLi'], df_final['Precio Mayoreo'], df_final['Descripción Truper']), start = 1):
                    screenshot_path = os.path.join(screenshot_folder, f'{mlm}.png')
                    if index == 1:
                        worksheet['A2'] = index
                        mlm_cell = worksheet['B2']
                        mlm_cell.value = mlm
                        mlm_cell.hyperlink = link
                        worksheet['D2'] = desc
                        worksheet['F2'] = precT
                        worksheet['H2'] = precML
                        worksheet['J2'] = hora
                    else:
                        worksheet[f'A{((index-1)*15)+(index*2)}'] = index
                        mlm_cell = worksheet[f'B{((index-1)*15)+(index*2)}']
                        mlm_cell.value = mlm
                        mlm_cell.hyperlink = link
                        worksheet[f'D{((index-1)*15)+(index*2)}'] = desc
                        worksheet[f'F{((index-1)*15)+(index*2)}'] = precT
                        worksheet[f'H{((index-1)*15)+(index*2)}'] = precML
                        worksheet[f'J{((index-1)*15)+(index*2)}'] = hora
                    
                    img = Image(screenshot_path)
                    img.width, img.height = 650, 300 
                    if index == 1:
                        worksheet.add_image(img, 'A3')
                    else:
                        worksheet.add_image(img, f'A{((index-1)*15)+(index*2)+1}')
                        
                for column in worksheet.columns:
                    column_letter = column[0].column_letter
                    for cell in column:                  
                        if column_letter in ['F']:
                            cell.number_format = '"$"#,##0.00'
                            
                        if column_letter in ['H']:
                            cell.number_format = '"$"#,##0.00'
                            
                        if column_letter in ['A']:
                            cell.alignment = Alignment(horizontal = 'center')
                            
                for cell in worksheet[1]:
                    cell.font = Font(color = 'FFFFFF', bold = False)
                    cell.fill = PatternFill(start_color = '000000', end_color = '000000', fill_type = 'solid')
                    cell.alignment = Alignment(horizontal = 'center')
                
                workbook.save(file_warnings)
                workbook.close()
                
                #Fin de ejecución
                end_time = time.time()
                elapsed = timedelta(seconds=int(end_time - start_time))
                
                dias = elapsed.days
                horas, resto = divmod(elapsed.seconds, 3600)
                minutos = resto // 60
                
                print('___________________________________________________________________')
                print(f'✅ Proceso finalizado, {len(df_final)} warnings fueron generados 🙂')
                print(f'⏳ El análisis de {len(df)} productos tardó: {dias} días, {horas} horas y {minutos} minutos')
        except Exception as e:
            print(f'Ocurrió un Error General: {e}')
            intento +=1
            referencia=i
            if i == 7312:
                referencia = 0
    print(f'⚠️ Ocurrieron {intento} intentos fallidos. Se suspende. ⚠️')  
    print(f'⭕🀄El último producto fue el que tiene el índice {i}🀄⭕')

##Inicialización
print('Cargando... Espere unos momentos....💿 ')
print('Iniciando Código Simple... ✅ ')
codigo_completo()  